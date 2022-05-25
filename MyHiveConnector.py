from pyhive import hive,presto
import requests
from pyhive.exc import *
from thrift.transport import TTransport
import time
from datetime import timedelta, date
from dateutil.relativedelta import relativedelta
from openpyxl import Workbook,styles
import pandas as pd
import login
from TCLIService.ttypes import TOperationState
from thrift.Thrift import TApplicationException
import re
class Presto(object):
	"""프레스토 커넥터"""
	def __init__(self):
		self._connect()
		self.field_names = None
	def _connect(self):
		self.conn = presto.connect(host = '10.220.151.103',
					port = 9080,
					protocol = 'http',
					requests_kwargs = {'auth':requests.auth.HTTPBasicAuth(login.ID,login.PASSWORD)})
		self.cursor = self.conn.cursor()
	def _reload(self):
		try:
			self.cursor.close()
			self.conn.close()
		except:
			pass
		self._connect()

	def sql(self,sql,retry = 1,verbose = True,debug = False):
		results = None
		ran = False
		for i in range(0,retry):
			try:
				stime= time.time()
				self.cursor.execute(sql)
				while True:
					ret = self.cursor.poll()

					if not ret:
						break
					else:
						stats  =ret['stats']
					time.sleep(0.2)
				if self.cursor.description is not None:
					results = self.cursor.fetchall()
					etime = time.time()
					qtime = etime -stime
					row_number=self.cursor.rownumber
					self.field_names = [field[0] for field in self.cursor.description]
				
				if verbose:
					try:
						row_number = stats['processedRows']
					except:
						row_number = 0

					qtime = time.time()-stime
					query =re.findall(r"(insert into .*)\s*|(create table .*)\s*|(drop table .*)\s*",sql.lower())
					
					if len(query) ==0:
						query  = str(sql[:50])
					else:
						query = ''.join(query[0])
					print(query+' ...')
					print("time taken: {:.2f}  row count : {}".format(qtime,row_number))
				ran = True
				break
			except KeyboardInterrupt:
				self.cursor.cancel()
				self._reload()
			except DatabaseError as e:
				e = e.args[0]
				print('----------------------------------------------------')
				print('[ERROR]{}\n {}'.format(time.ctime(),sql))
				print("ErrorCode:"+str(e['errorCode']))
				print("ErrorType:"+str(e['errorType']))
				print((e['message']))
				

				if e['errorCode'] == 131073 or e['errorType'] =='INSUFFICIENT_RESOURCES':
					print("하이브 에서 시도")
					break 
				if e['errorCode'] == 1 or e['errorType'] =='USER_ERROR"':
					break
				print('----------------------------------------------------')
		assert ran != False, "에러\n----------------------------------------------------"		
		return results



class Hive(object):
	"""하이브 커넥터
		LIST OF "USEFUL" METHODS
		-->sql
			쿼리수행 
		-->create_table
			테이블 생성
		-->to_excel
			엑셀 생성 (다중 시트 가능)
		-->to_csv
			CSV 생성
		-->to_sql
			Python 코드로 바꿔줌 (iteration 용)
	"""
	def __init__(self):
		self._connect()
		self.retry_time = 300
		self.lastsuccessfulsql = ''
		self.field_names = []
		self.DTYPE = ['string', 'tinyint', 'smallint','int','bigint','boolean','float','double','string','binary','timestamp']
		self.STORAGE_TYPE = ['orc','sequencefile','textfile']
	def _connect(self):
		self.conn = hive.connect(host = 'bdap.kt.com',
			port = 10000,
			auth = 'LDAP',
			username="10149348",
			password='''djsl0)qnrl''')
		self.cursor = self.conn.cursor()
	def _reload(self):
		try:
			self.cursor.close()
			self.conn.close()
		except:
			pass	
		self._connect()

	def sql(self,sql,retry = 1,debug = False,verbose = True):
		"""
			sql : query
			retry: return code X 뜰시 재시도 횟수
			verbose : 쿼리 수행 정보 
		"""
		#check connection is open
		if not self.conn._transport.isOpen():
			self._reload()
		sqlstate = None
		results = None
		is_option_chaged = False
		ran = False
		for i in range(0,retry):
			try:
				stime = time.time()
				self.cursor.execute(sql,_async = True)
				status = self.cursor.poll().operationState
				while status in (TOperationState.INITIALIZED_STATE, TOperationState.RUNNING_STATE):
					logs = self.cursor.fetch_logs()
					if debug:
						print('----------------------------------------------------')
						print('[STATUS]{}\n {}'.format(time.ctime(),sql))
						for message in logs:
							print(message)
						print('----------------------------------------------------')
					status = self.cursor.poll().operationState
				
				if status ==  (TOperationState.ERROR_STATE):
					raise OperationalError
				elif status == (TOperationState.FINISHED_STATE):
					logs = self.cursor.fetch_logs()
					if debug:
						for message in logs:
							print(message)
				else:
					for message in logs:
						print(message)
					print(status)
				if self.cursor.description is not None:
					results = self.cursor.fetchall()
					etime = time.time()
					qtime = etime -stime
					row_number=self.cursor.rownumber
					self.field_names = [field[0] for field in self.cursor.description]
				else:
					if verbose:
						try:
							row_number = re.findall(r"numRows=(\d+)",' '.join(logs))[0]
						except:
							row_number = 0

						qtime = time.time()-stime
						query =re.findall(r"(insert into .*)\s*|(create table .*)\s*|(drop table .*)\s*",sql.lower())
						
						if len(query) ==0:
							query  = str(sql[:50])
						else:
							query = ''.join(query[0])
						print(query+' ...')
						print("time taken: {:.2f}  row count : {}".format(qtime,row_number))					
				ran = True
				break
			except KeyboardInterrupt:
				self.cursor.cancel()
				self._reload()
			except TApplicationException as e:
				self._reload()
			except BrokenPipeError as e:
				self._reload()
			except OperationalError as e:
				print('----------------------------------------------------')
				print('[ERROR]{}\n {}'.format(time.ctime(),sql))
				try:
					print(e.args[0].status.errorMessage)
					sqlstate = e.args[0].status.sqlState
					print(sqlstate)
					if sqlstate !='08S01':
						break 
					else:
						is_option_chaged  = True 
						self.cursor.execute('set hive.auto.convert.join=False')
						time.sleep(self.retry_time)
						print("Retry Count {}".format(i))
						continue
				except:
					logs = self.cursor.fetch_logs()
					for message in logs:
						if message.startswith('ERROR'):
							print(message)
							code = re.findall(r"return code (\d)",message) 
							if len(code)>0:
								is_option_chaged= True
								self.cursor.execute('set hive.auto.convert.join=False')
								time.sleep(self.retry_time)
								print("Retry Count {}".format(i))
								continue
						if debug:
							print(message)

				
			except TTransport.TTransportException as e:
				print('[ERROR]{}\n {}'.format(time.ctime(),sql))
				print('Connection Closed')
				print((e))
				print('----------------------------------------------------')
				self._reload()
				print('Reconnected')
				print('----------------------------------------------------')
			except UnicodeDecodeError as e:
				i = i-1

			
		if is_option_chaged:
			self.cursor.execute('set hive.auto.convert.join=True')

		assert ran != False, "에러\n----------------------------------------------------"

		return results
	def _column_builder(self,columns):
		str_builder = []
		for column in columns.split(','):
			if len(column.split())==2:
				col_nm, col_type = column.split()
				if col_type.lower() not in self.DTYPE:
					raise AttributeError("컬럼타입 확인 : {}".format(col_type))
			elif len(column.split())==1:
				#set default type = string 
				col_nm =column.split()[0]
				col_type = 'string'
			else:
				raise AttributeError("컬럼명, 타입 확인 : {}".format(column))

			str_builder.append("{} {}".format(col_nm,col_type))
		
		col_str = ','.join(str_builder)
		return col_str
	
	def get_meta(self,table_name):
		"""
		Returns "DESC formated <table name>" in python dict format
		"""
		DESC_MAP_TABLE = {0:'COLUMN_INFO',1:'TABLE_INFO',2:'STORAGE_INFO'}
		t=self.sql('desc formatted {}'.format(table_name))
		idx = []
		for i in range(len(t)):
			if t[i][0].startswith('# col_name') or t[i][0].startswith('# Detailed Table') or t[i][0].startswith('# Storage Information'):
				idx.append(i)

		s_idx = [i+1 for i in idx]
		e_idx = [i-1 for i in idx if i>0]
		e_idx.append(-1)
		tmp=[]
		
		for x,y in zip(s_idx,e_idx):
			tmp.append(t[x:y])
		column_info = dict((str(x).strip().replace(':',''), str(y).strip())  if x!='' else (str(y).strip().replace(':',''),str(z).strip())  for x, y,z in tmp[0] )

		tmp2 = []
		for i in range(1,len(tmp)):
			tmp2+=tmp[i]
		table_info = dict((str(x).strip().replace(':',''), str(y).strip())  if x!='' else (str(y).strip().replace(':',''),str(z).strip())  for x, y,z in tmp2 )
		
		info = {}
		info['COLUMN'] = column_info
		info['TABLE'] = table_info
		return info 


	def create_table(self,table_name,columns = None,partitions = None,external = False,field_sep = None,line_sep= None,storage_type = 'ORC',drop  = False,compress = 'SNAPPY',debug = False,dry_run = False):
		"""
		table_name (필수): 스키마.테이블명 형식 
		columns (필수): 그냥 쉼표분리로.. 기본 dtype : string 
				columns = 'a,b,c,d' --> a,b,c,d 스트링 타입 컬럼
				columns = 'a,b,c int,d float' --> a,b, 스트링 c 인트 d 플롯 타입 
		partitions (옵션) : 파티션이 필요할때
		drop(옵션) 생성시 기존 테이블 드랍 
		dry_run(옵션): 생성 안하고 쿼리 반환 

		myhive.create_table('test.test',columns = 'a,b,c,d') --> test 스키마 test 테이블 컬럼들 a b c d 스트링타입 

		"""

		#stime= time.time()
		if len(table_name.split('.')) ==2:
			db = table_name.split('.')[0]
			tbl_nm = table_name.split('.')[1]
		else:
			print("테이블명 확인: {}".format(table_name))
			raise NameError(table_name)

		#스토리지 타입 확인
		if storage_type.lower() not in self.STORAGE_TYPE:
			raise AttributeError('미지원 스토리지 타입')

		if storage_type.lower() == 'orc' and compress.lower() in ('snappy','zlib'):
			compress_str = ',"orc.compress" = "{}"'.format(compress)
		else:
			compress_str = ''
		#컬럼 확인
		if columns is None or len(columns.split(',')) ==0:
			raise AttributeError("컬럼 없음")

		col_str = self._column_builder(columns)


		#파티션
		pts_str=''
		if partitions is not None:
			tmp_str = self._column_builder(partitions)
			pts_str = """PARTITIONED BY (
						{}
					)""".format(tmp_str)			

		#데이터 적재 시 
		ff_str = ''
		if field_sep is not None:
			ff_str = """row format delimited fields terminated by '{}' """.format(field_sep)
		lf_str = ''
		if line_sep is not None:
			lf_str = """lines terminated by '{}' """.format(line_sep)

		ext_str = ''
		if external:
			ext_str ='EXTERNAL'

		sql ="""CREATE {EXT} TABLE `{DB}`.`{TBL_NM}` (
					{COLS}
				)
				{PARTITIONS}
				{FIELD_FORMAT}
				{LINE_FORMAT}
				STORED AS {STORAGE_TYPE}
				TBLPROPERTIES (
					"ndap.table.storageType"="HDFS"
					{COMPRESS_TYPE}
				)""".format(DB = db,EXT = ext_str,TBL_NM = tbl_nm,COLS = col_str,PARTITIONS = pts_str,
					STORAGE_TYPE = storage_type,FIELD_FORMAT = ff_str,LINE_FORMAT=lf_str,COMPRESS_TYPE = compress_str)
		if drop:
			self.sql("drop table if exists {DB}.{TBL_NM} purge".format(DB = db,TBL_NM = tbl_nm))

		if not dry_run:
			self.sql(sql,debug = debug)
		else:
			print(sql)


	def  import_data(self,table_name,inpath =None,text_columns = None,columns = None,partitions = None,line_sep = '\n',field_sep = '\t',save_src = False,storage_type = 'TEXTFILE',drop = False):
		"""
		테이블 적재용 근데 잘 안씀
		"""
		def get_cols(columns):
			cols = []
			for column in columns.split(','):
				if len(column.split())==2:
					col_nm, col_type = column.split()
				elif len(column.split())==1:
					#set default type = string 
					col_nm =column.split()[0]
					col_type = 'string'
				cols.append(col_nm)
			return cols

		if inpath is None:
			raise AttributeError("파일주소 필수, inpath =")
		if text_columns is None:
			raise AttributeError("로드할 파일 컬럼 순서, text_columns =")

		if partitions is not None:
			if len(partitions.split(','))+len(columns.split(',')) != len(text_columns.split(',')):
				raise AttributeError("텍스트컬럼수 = 컬럼+파티션")
		else:
			if len(columns.split(',')) != len(text_columns.split(',')):
				raise AttributeError("텍스트컬럼수 = 컬럼+파티션")



		stime = time.time()

		#create 
		if storage_type.lower() != 'textfile':

			table_name_tmp =table_name+'_ctmp'
			self.create_table(table_name_tmp,columns = text_columns,
				external = save_src,line_sep = line_sep,
				field_sep = field_sep,storage_type='TEXTFILE',drop = drop)

			#load data 
			for f in inpath.split(','):
				sql = """load data inpath '{INPATH}' OVERWRITE into table {TMP_TBL}""".format(INPATH = f,TMP_TBL = table_name_tmp)
				self.sql(sql)
			
			pts_str = ''
			pts = ''
			cols = ','.join(get_cols(columns))
			if partitions is not None:
				pts_str = "PARTITION (" + ','.join(get_cols(partitions))+")"
				pts = ','.join(get_cols(partitions))
				cols = ','.join([cols,pts])
			
			
			#convert table storage
			self.create_table(table_name,columns =columns,partitions = partitions,storage_type = storage_type,drop = drop)
			#insert 
			sql = "insert into table {TBL_NM} {PARTITION} select {COLS} from {TMP_TBL_NM}".format(TBL_NM = table_name,PARTITION = pts_str,TMP_TBL_NM = table_name_tmp,COLS = cols)
			self.sql(sql)
			#drop temp
			sql = "drop table if exists {} purge".format(table_name_tmp)
			self.sql(sql)
			
			
		else:
			if partitions is not None:
				raise AttributeError("텍스트형식 파티션 미지원-> 파티션 storage_type = 'ORC' ")
			self.create_table(table_name,columns = columns,
							external = save_src,line_sep = line_sep,
							field_sep = field_sep,storage_type='TEXTFILE',drop = drop)

			#load data 
			for f in inpath.split(','):
				sql = """load data inpath '{INPATH}' OVERWRITE into table {TMP_TBL}""".format(INPATH = f,TMP_TBL = table_name)
				self.sql(sql)


		#get row cnt
		data = self.sql('select * from {} limit 5'.format(table_name))
		view =pd.DataFrame(data,columns = self.field_names)
		cnt = self.get_meta(table_name)['TABLE']['numRows']
		print("Import Job : {}, {} sec".format(table_name,round(time.time()-stime,0)))
		print("row count : {}".format(cnt))
		return view

		
	def to_excel(self,sql,wb = None,columns = None,sheet_name='Sheet',fname = str(int(time.time()))+'.xlsx',hold = False,border = True):
		"""
		sql(필수) : query 
		columns(옵션): 필드명 미지정시 bdap 기준 생성 
		sheet_name(옵션) : 시트명 
		fname(옵션) : 파일명, 미지정시 epoch time 기준 

		멀티시트 예제 

		wb = to_excel(sql,hold = True)
		wb = to_excel(sql,wb= wb,hold=True)
		to_excel(sql,wb= wb)


		"""
		data = self.sql(sql)
		if len(data)>1000000:
			print('[WARNING]{}\n {}'.format(time.ctime(),"데이터 수 100만줄  이상, to_csv 사용 "))
		if columns is None:
			column_names = self.field_names
		else:
			columns = columns.split(',')
			#trim 
			columns = [x.strip() for x in columns]
			if len(self.field_names) == len(columns):
				column_names = columns
			else:
				raise AttributeError("컬럼 길이 불일치")

		if not isinstance(wb,Workbook):		
			wb = Workbook()
			sheet = wb.active
			sheet.title = sheet_name
		else:
			sheet = wb.create_sheet(title = sheet_name)
		if border:
			thin_border = styles.borders.Border(left=styles.borders.Side(style='thin'),
										right=styles.borders.Side(style='thin'),
										top=styles.borders.Side(style='thin'),
										bottom=styles.borders.Side(style='thin'))
		
		for col,val in enumerate(column_names,start = 1):
			sheet.cell(row = 1,column=col).value = val
			if border:
				sheet.cell(row = 1,column=col).border = thin_border
		for row,vals in enumerate(data,start = 2):
			for col,val in enumerate(vals,start = 1):
				sheet.cell(row = row,column = col).value = val
				if border:
					sheet.cell(row = row,column=col).border = thin_border
		if hold:
			return wb
		else:
			wb.save(fname)

	def to_csv(self,sql,fname = str(int(time.time()))+'.csv',encoding = 'utf-8-sig',delimiter='\t', quoting = False,header = False):
		"""
		sql(필수)
		delimiter(옵션) 기본값 탭 
		quoting(옵션) 기본값 False
		header(옵션) 기본값 False 

		"""
		data = self.sql(sql)

		try:
			f = open('./'+fname,'w',encoding = encoding)
			if quoting:
				writer = csv.writer(f,delimiter = delimiter,quoting=csv.QUOTE_ALL)
			else:
				writer = csv.writer(f,delimiter = delimiter,quoting=csv.QUOTE_NONE)


			if header:
				writer.writerow(self.field_names)
			writer.writerows(data)

		except csv.Error as e:
			print("딜리미터 충돌?? --> quoting = True 또는  딜리미터 변경 ")
		except Exception as e :
			print(e)
	def to_sql(self,text,repl = None,retry = 3):
		"""
		text : BDAP 에서 돌아가는 쿼리들 ; 분리 그대로 
			ETL_YMD, BASE_YM, BASE_DATE 형식 감지 (오류 있을수 있음)
		retry: sql() 에서 재시도할 횟수 
		"""

		#\.[]{}()<>*+-=?^$|

		def format_builder(sql):
			if not repl:
				tmp = []
				if date_regex_var in sql:
					tmp.append(re.sub('\W+','',date_regex_var))
				if base_date_var in sql:
					tmp.append(re.sub('\W+','',base_date_var))
				if ym_regex_var in sql:
					tmp.append(re.sub('\W+','',ym_regex_var))
				if len(tmp)>0:
					return  sql+'.format('+', '.join(list(map(lambda x: x+' = '+x.lower(),tmp)))+')'
				else:
					return sql
			else:
				tmp =(set(re.findall("|".join(list(map(lambda x: "{}".format(x),vs))),sql)))
				if len(tmp)>0:
					return  sql+'.format('+', '.join(list(map(lambda x: x+' = '+x.lower(),tmp)))+')'
				else:
					return sql	
		def get_vars(sqls):
			tmp = []
			for sql in sqls:
				if date_regex_var in sql:
					tmp.append(re.sub('\W+','',date_regex_var).lower())
				if base_date_var in sql:
					tmp.append(re.sub('\W+','',base_date_var).lower())
				if ym_regex_var in sql:
					tmp.append(re.sub('\W+','',ym_regex_var).lower())
			return list(set(tmp))

		text_list = text.split(';')[0:-1]


		res = list(map(lambda x: re.sub(r"{",'{{',x.strip()),text_list))
		res = list(map(lambda x: re.sub(r"}",'}}',x.strip()),res))


		regex = r'\s{2,}'
		#res = list(map(lambda x: re.sub(regex,' ',x.strip()),res))

		comment = r"\s*(--.*)\s*"
		res = list(map(lambda x: re.sub(comment,'',x),res))
		res = [s for s in res if s!='']
		vs =[]
		if not repl:
			date_regex_var = '{ETL_YMD}'
			ym_regex_var = '{ETL_YM}'
			base_date_var = '{BASE_DATE}'


			date_regex = r'(?<![\d\*\+\-\/])[\d]{8}(?!(\d|(\s*\-)|(\s*\+)|(\s*\/)|(s*\*)))'
			res = list(map(lambda x: re.sub(date_regex,date_regex_var,x),res))
			basedate_regex = r'\d{4}-\d{2}-\d{2}'
			res = list(map(lambda x: re.sub(basedate_regex,base_date_var,x),res))
			ym_regex =  r'(?<![\d\*\+\-\/])[\d]{6}(?!(\d|(\s*\-)|(\s*\+)|(\s*\/)|(s*\*)))'
			res = list(map(lambda x: re.sub(ym_regex,ym_regex_var,x),res))
			res = list(map(lambda x: 'sql = """'+x+' """',res))
			vs = get_vars(res)
		else:
			for targets in repl.split('|'):
				if len(targets.split(':'))!=2:
					raise AttributeError("repl 입력 오류")
				target,var= targets.split(':')
				vs.append(var)
				regex = r'(?<![\d])[\d]{LEN}(?!(\d))'.format(LEN = '{'+str(len(target))+'}')
				res = list(map(lambda x: re.sub(regex,'{'+var+'}',x),res))
				#res = list(map(lambda x:x.replace(target,'{'+var+'}'),res))
			res = list(map(lambda x: 'sql = """'+x+' """',res))

		if retry ==1:
			execution = 'hive.sql(sql)'
		else:
			execution = 'hive.sql(sql,retry = {})'.format(retry)
		print("#파티션 쓰는 경우 repl 변수 활용, 예)  20190101:ETL_YMD|201901:ETL_YM")
		print("#==================Variables======================")
		for v in vs:
			print(v + ' =')

		print("#==================Python SQL======================")
		for line in list(map(lambda x: format_builder(x),res)):
			print(line)
			print(execution)
			print()


	def  set_retry_time(self,sleep_time):
		"""
		재시도 대기 시간 설정용 
		잘 안씀 
		"""
		self.retry_time= sleep_time
	def _doc(self):
		print("FUNCTION LIST")
		print("""============================================================""")
		doc = """sql(sql,retry = 1,debug = False)"""
		print(doc)
		print("""셀렉트 쿼리시 data 반환""")
		print("""------------------------------------------------------------""")
		print("""============================================================""")
		doc = """create_table(table_name,columns = None,partitions = None,storage_type = 'ORC',drop  = False)\n"""
		doc+="-필수 컬럼\n"
		doc+="\ttable_name : 테이블 명\n"
		doc+="\tcolumns : 컬럼 명\n"
		doc+="\tpartitions: 파티션 명\n"
		doc+="-기타 옵션들 \n"
		doc+="\tstorage_type : 하이브 저장 방식   \n"
		doc+="\tdrop : 생성 전 드랍 여부, 기본값 False\n"
		doc+="-기타사항 \n"
		doc+="컬럼과 파티션 입력 형식 svc_cont_id string,rqt_host_adr string\n 또는 svc_cont_id,rqt_host_adr = svc_cont_id string, rqt_host_adr string"
		print(doc)
		doc =''
		print("""============================================================""")
		print("""import_data(table_name,inpath =None,text_columns = None,columns = None,partitions = None,line_sep = '\\n',field_sep = '\\t',storage_type = 'TEXTFILE',drop = False)""")
		print("""------------------------------------------------------------""")
		doc+="-필수 컬럼\n"
		doc+="\ttable_name : 테이블 명\n"
		doc+="\tinpath :  =파일 경로\n"
		doc+="\tcolumns : 컬럼 명\n"
		doc+="-기타 옵션들 \n"
		doc+="\tpartitions: 파티션 명\n"
		doc+="\tline_sep : 라인분리 구분자 기본값 \\n \n"
		doc+="\tfield_sep : 컬럼 구분자 기본값 \\t \n"
		doc+="\tstorage_type : 하이브 저장 방식   \n"
		doc+="\tdrop : 생성 전 드랍 여부, 기본값 False\n"
		doc+="-기타사항 \n"
		doc+="컬럼과 파티션 입력 형식 svc_cont_id string,rqt_host_adr string\n 또는 svc_cont_id,rqt_host_adr = svc_cont_id string, rqt_host_adr string"
		print(doc)
		doc =''
		print("""============================================================""")
		print("""to_excel(sql,wb = None,sheet_name='Sheet',fname = str(int(time.time()))+'.xlsx',hold = False,border = True)""")
		print("""------------------------------------------------------------""")
		doc+="-필수 컬럼\n"
		doc+="\tsql : 쿼리\n"
		doc+="-기타 옵션들 \n"
		doc+="\twb:: 엑셀 오브젝트 (여러 시트 생성시 사용) \n"
		doc+="\tsheet_name: 시트명 \n"
		doc+="\tfname : 파일명, 기본값 unix timestamp  \n"
		doc+="\thold : 저장하지 않고 wb 오브젝트 반환   \n"
		doc+="\tborder : 테두리  \n"
		doc+="-기타사항 \n"
		doc+="wb = to_excel(sql,hold = True), to_excel(sql2,wb = wb) --> 동파일 시트2개 생성 "
		print(doc)
		doc =''
		print("""============================================================""")
		print("""to_csv(self,sql,fname = str(int(time.time()))+'.csv',encoding = 'utf-8-sig',delimiter='\t', quoting = False,header = False)""")
		print("""------------------------------------------------------------""")
		doc+="-필수 컬럼\n"
		doc+="\tsql : 쿼리\n"
		doc+="-기타 옵션들 \n"
		doc+="\tfname : 파일명, 기본값-unix timestamp  \n"
		doc+="\tencoding : 인코딩, 기본값, bdap과 도일  \n"
		doc+="\tdelimiter : 구분자, 기본값 탭(\\t)  \n"
		doc+="\tquoting : 컬럼 따옴표  여부 , 기본값 : 감싸지 않기  \n"
		doc+="\theader : 헤더 추가여부  \n"
		print(doc)
		doc =''
		print("""============================================================""")
		print("""get_meta(table_name)""")
		print("""------------------------------------------------------------""")
		doc+="-필수 컬럼\n"
		doc+="\ttable_name : 테이블명 \n"
		doc+="-기타사항 \n"
		doc+="get_meta(table_name)['COLUMN'] --> 컬럼 정보 반환 None 무시\n"
		doc+="get_meta(table_name)['TABLE'] --> 테이블 정보 ex) row 수 등 "
		print(doc)
		doc =''
		print("""============================================================""")